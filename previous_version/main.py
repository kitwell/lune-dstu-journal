import PySimpleGUI as sg
import openpyxl.styles
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from copy import copy
from re import search, split, fullmatch, match
from pathlib import Path


def save_font_settings(c_type, c_family, c_style, c_size):
    settings['FONT']['current_font_type'] = c_type
    settings['FONT']['current_font_family'] = c_family
    settings['FONT']['current_font_style'] = c_style
    settings['FONT']['current_font_size'] = c_size


def quick_message():
    layout = [
        [sg.Text('Цвет по умолчанию расположен в правом нижнем углу')],
        [sg.Push(), sg.Button('OК', key='-OK-'),
         sg.Checkbox(text='не показывать больше', key='-SHOW_AGAIN-', enable_events=True), sg.Push()]
    ]
    window = sg.Window(title='', no_titlebar=True, layout=layout, keep_on_top=True, modal=True)
    while True:
        event, values = window.read()
        if event in (sg.WINDOW_CLOSED, '-OK-'):
            break
        if event == '-SHOW_AGAIN-':
            settings['GUI']['show_quick_message'] = not values['-SHOW_AGAIN-']
    window.close()


def color_map(default_color):
    colors = (
        ('#8B0000', '#8B4500', '#8B6914', '#008B45', '#104E8B', '#00008B', '#68228B', '#000000', '#7D7D7D'),
        ('#CD0000', '#CD6600', '#CD9B1D', '#00CD00', '#1874CD', '#0000CD', '#9932CC', '#141414', '#919191'),
        ('#EE0000', '#EE7600', '#DAA520', '#00EE00', '#1C86EE', '#0000EE', '#B23AEE', '#1A1A1A', '#A6A6A6'),
        ('#FF0000', '#FF7F24', '#EEB422', '#00FF00', '#1E90FF', '#0000FF', '#BF3EFF', '#262626', '#BABABA'),
        ('#FF2400', '#FF7F00', '#FFD700', '#7FFF00', '#42AAFF', '#2A52BE', '#D15FEE', '#2E2E2E', '#CFCFCF'),
        ('#EE2C2C', '#FF8C00', '#EEEE00', '#ADFF2F', '#00B2EE', '#4169E1', '#E066FF', '#383838', '#E3E3E3'),
        ('#FF3030', '#EE9A00', '#FFFF00', '#C0FF3E', '#00CDCD', '#1F75FE', '#EE7AE9', '#454545', '#F7F7F7'),
        ('#EE3B3B', '#FFA500', '#FFFF66', '#BCEE68', '#00EEEE', '#008CF0', '#FF83FA', '#616161', '#FFFFFF'),
        ('#FF4040', '#FFB02E', '#FFF68F', '#CAFF70', '#00FFFF', '#1FAEE9', '#EEAEEE', '#6B6B6B')
    )

    layout = []
    for row in colors:
        layout_row = []
        for color in row:
            layout_row.append(sg.Radio(text='     ', font=('Any', 17), key=color, background_color=color,
                                       group_id=0, circle_color=color, enable_events=True))
        layout.append(layout_row)
    layout[-1].append(sg.Radio(text='     ', font=('Any', 17), key=default_color, background_color=default_color, group_id=0,
                               enable_events=True, circle_color=default_color, default=True))
    layout.append(
        [sg.Input(default_text=default_color, size=9, text_color=default_color,
                  key='-HEX_INPUT-', font=(settings['FONT']['font_family'], settings['FONT']['font_size'], 'bold'), justification='center'),
         sg.Button('Ок', key='-HEX_OK-'), sg.Button('Отменить', key='-HEX_CANCEL-')])
    window = sg.Window(title='', layout=layout, no_titlebar=True, element_padding=((3, 3), (3, 3)), grab_anywhere=True,
                       use_default_focus=False, finalize=True, margins=(5, 5),
                       border_depth=4, element_justification='center', keep_on_top=True)
    for key in list(window.key_dict.keys())[:-3]:
        window[key].widget.configure(indicatoron=False)
    if settings['GUI']['show_quick_message']:
        quick_message()
    while True:
        event, values = window.read()
        if event == '-HEX_CANCEL-':
            break
        if event == '-HEX_OK-':
            sg.Popup(sg.Button('fff'))
        if event.startswith('#'):
            window['-HEX_INPUT-'].update(event)
            window['-HEX_INPUT-'].update(text_color=event)
    window.close()


def show_subgroup(check, sbjct, act, sub):
    if check:
        return f'{sbjct} ({act}, {sub})'
    else:
        return f'{sbjct} ({act})'


def is_valid_path(filepath, location):  # функция для проверки пути файла
    filepaths = filepath.split(';')
    if all(filepaths):
        for item in filepaths:
            if Path(item).exists():
                return True
    # print(filepath)
    # sg.popup_animated(image_source='images/loading.gif', time_between_frames=5000, transparent_color='black')
    sg.popup_no_titlebar('Путь некорректен!',
                         relative_location=(location[0] - 660, location[1] - 440))


def copy_cell(source_cell, target_cell):  # функция для копирования формата ячеек
    target_cell.alignment = copy(source_cell.alignment)
    target_cell.value = copy(source_cell.value)
    target_cell.font = copy(source_cell.font)
    target_cell.border = copy(source_cell.border)
    target_cell.fill = copy(source_cell.fill)
    target_cell.number_format = copy(source_cell.number_format)
    target_cell.protection = copy(source_cell.protection)


def make_shorter(name, difference, symbols, lower_cons):  # функция для сокращения названий предметов
    name_length = len(name)
    short_name = split(r'(\W+)', name)
    for index, word in enumerate(short_name):
        if not fullmatch(r'\W+', word) and word:
            if name_length - len(''.join(short_name)) < difference:
                stop = symbols
                while word[:stop][-1].lower() not in lower_cons:
                    if stop >= len(word) - 1:
                        break
                    stop += 1
                else:
                    if stop < len(word) - 1:
                        short_name[index] = word[:stop] + '.'
            else:
                return ''.join(short_name).replace('..', '.')
    joined_short_name = ''.join(short_name).replace('..', '.')
    return make_shorter(joined_short_name, difference - (name_length - len(joined_short_name)), symbols - 1, lower_cons)


def main_window():
    font_size = settings['FONT']['font_size']
    # font_color = settings['FONT']['font_color']
    font_family = settings['FONT']['font_family']
    font_style = settings['FONT']['font_style']
    path_font_size = settings['FONT']['path_font_size']
    # path_font_color = settings['FONT']['path_font_color']
    path_font_family = settings['FONT']['path_font_family']
    path_font_style = settings['FONT']['path_font_style']
    tab_font_size = settings['FONT']['tab_font_size']
    # tab_font_color = settings['FONT']['tab_font_color']
    tab_font_family = settings['FONT']['tab_font_family']
    tab_font_style = settings['FONT']['tab_font_style']
    current_font_type = settings['FONT']['current_font_type']
    current_font_family = settings['FONT']['current_font_family']
    current_font_style = settings['FONT']['current_font_style']
    current_font_size = settings['FONT']['current_font_size']
    if settings['GUI']['hide_tooltips']:
        tooltip_time = 999999999
    else:
        tooltip_time = 100
    sg.set_options(font=(font_family, font_size, font_style),
                   tooltip_time=tooltip_time, tooltip_font=(font_family, int(int(font_size) / 1.2)))
    main_window_location = eval(settings['GUI']['window_coordinates'])
    theme = settings['GUI']['theme']
    sg.theme(themes_dict[theme])
    theme_color = sg.theme_text_color().upper()
    tab_vertical_location = settings['GUI']['tab_vertical_location']
    tab_horizontal_location = settings['GUI']['tab_horizontal_location']
    # print(sg.theme_text_color().upper())
    # themes = ('дикий синий', 'васильковое поле', 'песчаная буря', 'лунное затмение', 'дождливый лес', 'DarkBlue17',
    #           'DarkGreen', 'DarkGrey5', 'DarkTeal', 'DarkTeal1', 'DefaultNoMoreNagging', 'DarkTeal7',
    #           'DarkTeal8', 'GreenMono', 'LightGreen4', 'Material1', 'Purple', 'Reddit',
    #           'SandyBeach', 'TanBlue')
    consonants = 'бвгджзйклмнпрстфхцчшщ'
    main_tab = [
        [sg.VPush()],
        [sg.I(key='-INPUT_PATH-', default_text=settings['EXCEL']['input_filepath'],
              font=(path_font_family, path_font_size, path_font_style), text_color=theme_color, expand_x=True),
         sg.FilesBrowse(file_types=(('Excel Files', '*.xlsx'),), button_text='Обзор', key='-VIEW-')],
        [sg.I(key='-OUTPUT_PATH-', default_text=settings['EXCEL']['output_filepath'],
              font=(path_font_family, path_font_size, path_font_style), text_color=theme_color, expand_x=True),
         sg.FolderBrowse(button_text='Обзор')],
        [sg.B(button_text='Выполнить', key='-TAKE_NAMES-'),
         # sg.Checkbox(text='бустануть', default=settings['EXCEL']['boosted'], key='-BOOSTED-',
         #             tooltip=settings['GUI']['boost_tooltip']),
         # sg.Checkbox(text='подгруппы', default=settings['EXCEL']['show_subgroup'], key='-SHOW_SUBGROUP-',
         #             tooltip=settings['GUI']['subgroup_tooltip']),
         sg.Button(button_text='Отменить', visible=False, key='-CANCEL-'),
         sg.ProgressBar(0, orientation='h', k='-PBAR-', expand_x=True, visible=False, s=(25, int(int(font_size) * 1.5)))],
        # [sg.ProgressBar(0, orientation='h', k='-PBAR-', expand_x=True, visible=False, s=(25, 25))],
        [sg.VPush()]
    ]

    settings_tab = [
        [sg.VPush()],
        [sg.Text('Имя сохранённого файла', relief='groove', border_width=5), sg.Input(size=(41, 1), key='-FILE_NAME-', enable_events=True, default_text=settings['EXCEL']['output_file_name'], justification='center')],
        [sg.Text('Элементы меню', relief='groove', border_width=5), sg.Combo(('сверху', 'снизу'), default_value=tab_vertical_location, readonly=True, size=7, enable_events=True, key='-VERTICAL-'),
         sg.Combo(('слева', 'посередине', 'справа'), default_value=tab_horizontal_location, readonly=True, size=11, enable_events=True, key='-HORIZONTAL-'),
         sg.Push(),
         sg.Text('Тема', relief='groove', border_width=5), sg.Combo(themes, default_value=theme, readonly=True, enable_events=True, size=17, key='-THEME-')],
        [sg.Push(), sg.Combo(values=('Основной текст', 'Текст путей', 'Элементы меню'), default_value=current_font_type, readonly=True, size=16, enable_events=True, key='-FONT_TYPE-'),
         sg.Combo(values=('по умолчанию', 'Courier New', 'Times New Roman', 'Comic Sans MS'), default_value=current_font_family, readonly=True, size=16, enable_events=True, key='-FONT_FAMILY-'),
         sg.Combo(values=('обычный', 'курсив', 'жирный', 'подчёркнутый'), default_value=current_font_style, readonly=True, size=14, enable_events=True, key='-FONT_STYLE-'),
         sg.Spin(values=tuple(range(14, 31)), initial_value=current_font_size, size=2, enable_events=True, key='-FONT_SIZE-', bind_return_key=True),
         sg.Button(button_color=theme_color, mouseover_colors=theme_color, button_text='    ', border_width=4,
                   key='-FONT_COLOR-', font=(font_family, font_size, 'normal')), sg.Push()],
         # sg.Button(image_source='images/colorwheel.png', image_subsample=subsamples_dict[settings['FONT']['font_size']], key='-FONT_COLOR-'), sg.Push()],

        [sg.Push(), sg.Checkbox(text='буст', default=settings['EXCEL']['boosted'], key='-BOOSTED-',
                     tooltip='Ускорить выполнение, игнорируя пустые ячейки', enable_events=True),
         sg.Checkbox(text='подгруппа', default=settings['EXCEL']['show_subgroup'], key='-SHOW_SUBGROUP-',
                     tooltip='Отображать номер подгруппы для лабораторных занятий', enable_events=True),
         sg.Checkbox(text='пропуски', default=settings['EXCEL']['show_misses'], key='-SHOW_MISSES-',
                     tooltip='Добавить лист с общим числом пропусков по каждому студенту', enable_events=True),
         sg.Checkbox(text='курсор+', default=settings['GUI']['custom_cursors'], key='-CURSOR-',
                     tooltip='Добавить разнообразия', enable_events=True),
         sg.Checkbox(text='отключить подсказки', default=settings['GUI']['hide_tooltips'], key='-TOOLTIPS-',
                     enable_events=True), sg.Push()],
    ]

    # if tab_horizontal_location.strip() == 'слева':
    #     settings_tab[-1] = settings_tab[-1] + [sg.Push()]
    #     settings_tab[-2] = settings_tab[-2] + [sg.Push()]
    # elif tab_horizontal_location.strip() == 'посередине':
    #     settings_tab[-1] = [sg.Push()] + settings_tab[-1] + [sg.Push()]
    #     settings_tab[-2] = [sg.Push()] + settings_tab[-2] + [sg.Push()]
    # else:
    #     settings_tab[-1] = [sg.Push()] + settings_tab[-1]
    #     settings_tab[-2] = [sg.Push()] + settings_tab[-2]

    about_tab = [

    ]
    layout = [
        [sg.TabGroup([
            [sg.Tab(' Главная ', main_tab, key='-MAIN_TAB-'),
             sg.Tab(' Настройки ', settings_tab, key='-SETTINGS_TAB-'),
             sg.Tab(' О программе ', about_tab, key='-ABOUT_TAB-')]],
            font=(tab_font_family, tab_font_size, tab_font_style),
            tab_border_width=0, border_width=0, tab_location=tab_locations[f'{tab_vertical_location} {tab_horizontal_location}'], focus_color='',
            pad=((10, 10), (5, 5)), enable_events=True, key='-SWITCH_TAB-', title_color=theme_color)]
    ]

    window = sg.Window(title=settings['GUI']['title'], layout=layout, use_default_focus=False,
                       location=main_window_location, use_custom_titlebar=True, enable_close_attempted_event=True,
                       finalize=True, return_keyboard_events=True, keep_on_top=True)

    tabs_keys = ('-MAIN_TAB-', '-SETTINGS_TAB-', '-ABOUT_TAB-')
    window[settings['GUI']['current_tab']].select()
    window['__TITLEBAR MAXIMIZE__'].update(visible=False)
    window.set_cursor('hand2')
    # здесь функция для настройки всего-всего перед запуском
    save = False
    # print(list(window.key_dict.keys()))
    while True:
        event, values = window.read()
        if save:
            save_font_settings(values['-FONT_TYPE-'], values['-FONT_FAMILY-'], values['-FONT_STYLE-'], values['-FONT_SIZE-'])
            save = False
        if event == sg.WINDOW_CLOSE_ATTEMPTED_EVENT:
            settings['GUI']['window_coordinates'] = str(window.current_location())
            # save_font_settings(values['-FONT_TYPE-'], values['-FONT_FAMILY-'], values['-FONT_STYLE-'])
            break
        if event == sg.WINDOW_CLOSED:
            break
        # if event:
        #     window.force_focus()
        if event == '-FONT_SIZE-':
            font_size_value = int(values['-FONT_SIZE-'])
            # print(type(font_size_value))
            if font_size_value > 30:
                font_size_value = 30
            elif font_size_value < 14:
                font_size_value = 14
            elif font_size_value == int(settings['FONT'][f'{font_types[values['-FONT_TYPE-']]}_size']):
                # print(type(font_size_value))
                continue
            # font_type_value = values['-FONT_TYPE-']
            settings['FONT'][f'{font_types[values['-FONT_TYPE-']]}_size'] = font_size_value
            settings['FONT']['current_font_size'] = font_size_value
            settings['GUI']['window_coordinates'] = str(window.current_location())
            window.close()
            main_window()
            # elif font_type_value == 'Текст путей':
            #     window['-INPUT_PATH-'].update(font=(path_font_family, font_size_value, path_font_style))
            #     window['-OUTPUT_PATH-'].Update(font=(path_font_family, font_size_value, path_font_style))
            #     window['-INPUT_PATH-'].update()
            # else:
            #     window['-SWITCH_TAB-'].update(font=(tab_font_family, font_size_value, tab_font_style))
                # window['-SETTINGS_TAB-'].update(font=(tab_font_family, font_size_value, tab_font_style))
                # window['-ABOUT_TAB-'].update(font=(tab_font_family, font_size_value, tab_font_style))
            # window['-VIEW-'].update(font=('default', values['-FONT_SIZE-'], 'normal'))
        if event == '-FONT_TYPE-':  # обновление текущих значений полей, касающихся выбранного шрифта font_type
            font_type_value = values['-FONT_TYPE-']
            window['-FONT_FAMILY-'].update(value=settings['FONT'][f'{font_types[font_type_value]}_family'])
            window['-FONT_STYLE-'].update(value=list(font_styles.keys())[list(font_styles.values()).index(settings['FONT'][f'{font_types[font_type_value]}_style'].split()[-1])])
            window['-FONT_SIZE-'].update(value=settings['FONT'][f'{font_types[font_type_value]}_size'])
            # settings['FONT']['current_font_type'] = values['-FONT_TYPE-']
            save = True
            continue
        if event == '-FONT_STYLE-':  # обновление стиля выбранного шрифта font_type
            font_style_value = font_styles[values['-FONT_STYLE-']]
            settings_font_style = settings['FONT'][f'{font_types[values['-FONT_TYPE-']]}_style']
            if font_style_value != 'normal':
                if font_style_value in settings_font_style:
                    settings_font_style = settings_font_style.replace(font_style_value, '')
                    settings['FONT'][f'{font_types[values['-FONT_TYPE-']]}_style'] = settings_font_style.rstrip()
                else:
                    settings['FONT'][f'{font_types[values['-FONT_TYPE-']]}_style'] = settings_font_style + ' ' + font_style_value
            else:
                settings['FONT'][f'{font_types[values['-FONT_TYPE-']]}_style'] = 'normal'

            # window['-FONT_STYLE-'].update(list(font_styles.keys())[list(font_styles.values()).index(
            #     settings['FONT'][f'{font_types[values['-FONT_TYPE-']]}_style'].split()[-1])])
            # save_font_settings(values['-FONT_TYPE-'], values['-FONT_FAMILY-'], values['-FONT_STYLE-'])
            # settings['FONT']['current_font_style'] = values['-FONT_STYLE-']

            settings['FONT']['current_font_style'] = list(font_styles.keys())[list(font_styles.values()).index(
                settings['FONT'][f'{font_types[values['-FONT_TYPE-']]}_style'].split()[-1])]
            settings['GUI']['window_coordinates'] = str(window.current_location())
            window.close()
            main_window()
        if event == '-FONT_FAMILY-':
            font_family_value = values['-FONT_FAMILY-']
            if font_family_value == settings['FONT'][f'{font_types[values['-FONT_TYPE-']]}_family']:
                continue
            settings['FONT'][f'{font_types[values['-FONT_TYPE-']]}_family'] = font_family_value

            # save_font_settings(values['-FONT_TYPE-'], values['-FONT_FAMILY-'], values['-FONT_STYLE-'])

            settings['FONT']['current_font_family'] = font_family_value
            settings['GUI']['window_coordinates'] = str(window.current_location())
            window.close()
            main_window()
        if event == '-VERTICAL-':
            vertical_value = values['-VERTICAL-']
            if vertical_value == settings['GUI']['tab_vertical_location']:
                continue
            settings['GUI']['tab_vertical_location'] = vertical_value
            settings['GUI']['window_coordinates'] = str(window.current_location())
            # save_font_settings(values['-FONT_TYPE-'], values['-FONT_FAMILY-'], values['-FONT_STYLE-'])
            window.close()
            main_window()
        if event == '-HORIZONTAL-':
            horizontal_value = values['-HORIZONTAL-']
            if horizontal_value == settings['GUI']['tab_horizontal_location']:
                continue
            settings['GUI']['tab_horizontal_location'] = horizontal_value
            settings['GUI']['window_coordinates'] = str(window.current_location())
            # save_font_settings(values['-FONT_TYPE-'], values['-FONT_FAMILY-'], values['-FONT_STYLE-'])
            window.close()
            main_window()
        if event == '-TOOLTIPS-':
            settings['GUI']['hide_tooltips'] = values['-TOOLTIPS-']
            settings['GUI']['window_coordinates'] = str(window.current_location())
            # save_font_settings(values['-FONT_TYPE-'], values['-FONT_FAMILY-'], values['-FONT_STYLE-'])
            window.close()
            main_window()
        if event == '-SHOW_SUBGROUP-':
            settings['EXCEL']['show_subgroup'] = values['-SHOW_SUBGROUP-']
        if event == '-SHOW_MISSES-':
            settings['EXCEL']['show_misses'] = values['-SHOW_MISSES-']
        if event == '-BOOSTED-':
            settings['EXCEL']['boosted'] = values['-BOOSTED-']
        if event == '-FILE_NAME-':
            settings['EXCEL']['output_file_name'] = values['-FILE_NAME-']
        if event == '-FONT_COLOR-':
            color_map(theme_color)
        if event == '-THEME-':
            theme_value = values['-THEME-']
            if theme_value == settings['GUI']['theme']:
                continue
            # print('ff')
            settings['GUI']['theme'] = theme_value
            settings['GUI']['window_coordinates'] = str(window.current_location())
            # save_font_settings(values['-FONT_TYPE-'], values['-FONT_FAMILY-'], values['-FONT_STYLE-'])
            window.close()
            main_window()
        if event == '-SWITCH_TAB-':
            window.force_focus()
            # print(window['-TITLE-'].fonts_installed_list())
        if event in ('F1:112', 'F2:113', 'F3:114'):
            window[tabs_keys[int(event[-1]) - 2]].select()

        # if event:
        #     print(event)
        # if event in ('left arrow', 'RIGHT', 'up'):
        #     window[tabs_keys[int(event) - 1]].select()
        # elif event == '-INPUT_BROWSE-':
        #     window['-OUTPUT_PATH-'].update(values['-INPUT_BROWSE-'].split(';')[0])
        if (event == '-TAKE_NAMES-' and is_valid_path(values['-INPUT_PATH-'], main_window_location) and
                is_valid_path(values['-OUTPUT_PATH-'], main_window_location)):
            final_workbook = Workbook()
            group = ''
            flag = True
            window.set_cursor('watch')
            students = dict()
            for count, path in enumerate(values['-INPUT_PATH-'].split(';')):  # цикл для прохода по каждой книге
                event, values = window.read(timeout=0)
                if event == '-CANCEL-':
                    flag = False
                    # window['-PBAR-'].update_bar(len(values['-INPUT_PATH-'].split(';')) - 1,
                    #                             len(values['-INPUT_PATH-'].split(';')) - 1)
                    # window['-CANCEL-'].update(visible=False)
                    break
                window['-CANCEL-'].update(visible=True)
                window['-PBAR-'].update(visible=True)
                window['-PBAR-'].update_bar(current_count=count, max=len(values['-INPUT_PATH-'].split(';')) - 1)
                active_workbook = load_workbook(path)
                active_sheet = active_workbook.active
                book_title = split(r'[\\/]', path)[-1]
                subject_info = search(r'\((?P<subject_name>.+?)\)', book_title)  # список с информацией о предмете
                # location = window.current_location()
                if subject_info:
                    subject_info = [item.strip() for item in split(r'[.,;]', subject_info.group('subject_name'))]
                    # print(subject_info)
                    if len(subject_info) == 4:
                        subject_name, subgroup, activity, group = subject_info
                        sheet_name = show_subgroup(values['-SHOW_SUBGROUP-'], subject_name, activity, subgroup)
                        if len(sheet_name) > 31:
                            subject_name = make_shorter(subject_name, len(sheet_name) - 31, len(subject_name),
                                                        consonants)
                            sheet_name = show_subgroup(values['-SHOW_SUBGROUP-'], subject_name, activity, subgroup)
                    else:
                        subject_name, activity, group = subject_info
                        sheet_name = f'{subject_name} ({activity})'
                        if len(sheet_name) > 31:
                            subject_name = make_shorter(subject_name, len(sheet_name) - 31, len(subject_name),
                                                        consonants)
                            sheet_name = f'{subject_name} ({activity})'
                    # print(sheet_name)
                    final_sheet = final_workbook.create_sheet(sheet_name)
                else:
                    final_sheet = final_workbook.create_sheet(str(count+1))

                top_left_cells = {str(merged_range).split(':')[0]:  # словарь со всеми диапазонами объединённых ячеек
                                  str(merged_range) for merged_range in active_sheet.merged_cells.ranges}

                for item in active_sheet[active_sheet.max_row + 100]:  # копирование ширины каждого столбца
                    final_sheet.column_dimensions[item.column_letter].width = \
                        active_sheet.column_dimensions[item.column_letter].width

                for row in active_sheet.iter_rows():  # копирование значений ячеек построчно
                    student_name = ''
                    misses = 0
                    for cell in row:
                        if values['-BOOSTED-']:
                            cells_interval = top_left_cells.get(cell.coordinate, False)
                            if cells_interval:
                                final_sheet.merge_cells(cells_interval)
                                copy_cell(cell, final_sheet[cell.coordinate])
                            elif cell.value is not None:
                                if match(r'[A-zА-яёЁ]{2,}\s+[A-ZА-ЯЁ]\.\s*[A-ZА-ЯЁ]\.', str(cell.value)):
                                    student_name = cell.value
                                elif '%' in str(cell.value):
                                    misses += int(cell.value.split()[0])
                                copy_cell(cell, final_sheet[cell.coordinate])
                            else:
                                continue
                        else:
                            if cell.has_style:
                                cells_interval = top_left_cells.get(cell.coordinate, False)
                                if cells_interval:
                                    final_sheet.merge_cells(cells_interval)
                                copy_cell(cell, final_sheet[cell.coordinate])
                    if student_name:
                        students[student_name] = students.get(student_name, 0) + misses
                final_sheet.freeze_panes = active_sheet.freeze_panes  # копирование фиксированных строк, если имеются
            window['-CANCEL-'].update(visible=False)
            window['-PBAR-'].update(visible=False)
            # window['-PBAR-'].Widget.master.pack_forget()
            # window['-PBAR-'].Widget.master.pack()
            # window['-PBAR-'].Widget.master.config(max_value=0, orientation='h', s=(20, 20), k='-PBAR-', expand_x=True)
            # window.minimize()
            final_workbook.remove(final_workbook.worksheets[0])  # удаление первого листа, созданного по умолчанию

            if values['-SHOW_MISSES-']:  # добавляем информацию о пропусках за семестр на дополнительный лист
                additional_sheet = final_workbook.create_sheet("Статистика по пропускам")
                additional_sheet['A1'] = 'ФИО обучающегося'
                additional_sheet['A1'].font = Font(bold=True)
                additional_sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
                additional_sheet['B1'] = 'Пропущено часов'
                additional_sheet['B1'].font = Font(bold=True)
                additional_sheet['B1'].alignment = Alignment(horizontal='center', vertical='center')
                studs_and_misses = tuple(students.items())
                for i in range(2, len(studs_and_misses) + 2):
                    additional_sheet[f'A{i}'], additional_sheet[f'B{i}'] = studs_and_misses[i - 2]
                    additional_sheet[f'B{i}'].alignment = Alignment(horizontal='center')
                    # if studs_and_misses[i - 2][-1] >= 60:
                    #     additional_sheet[f'A{i}'].fill = openpyxl.styles.PatternFill(fgColor='ffee58')
                    #     additional_sheet[f'B{i}'].fill = openpyxl.styles.PatternFill(fgColor='ffee58')
                additional_sheet.column_dimensions['A'].width = len(additional_sheet['A1'].value) * 1.3
                additional_sheet.column_dimensions['B'].width = len(additional_sheet['B1'].value) * 1.3

            file_name = settings['EXCEL']['output_file_name']
            if flag:
                for path in values['-OUTPUT_PATH-'].split(';'):
                    try:
                        if file_name:
                            final_workbook.save(fr'{path}/{file_name}.xlsx')  # сохранение книги
                        else:
                            final_workbook.save(fr'{path}/Журнал {group}.xlsx')  # сохранение книги
                    except PermissionError:
                        sg.popup_no_titlebar(f'У вас нет доступа к директории {path} или файл используется',
                                             relative_location=(
                                                 main_window_location[0] - 660, main_window_location[1] - 440))
                    except OSError:
                        sg.popup_no_titlebar(f'Имя файла ({file_name}) содержит запрещённые символы',
                                             relative_location=(
                                                 main_window_location[0] - 660, main_window_location[1] - 440))

            window.set_cursor('arrow')
    # здесь функция для записи настроек с использованием словарей
    window.close()


if __name__ == '__main__':
    SETTINGS_PATH = str(Path.cwd())
    settings = sg.UserSettings(
        path=SETTINGS_PATH, filename='config - Copy.ini', use_config_file=True, convert_bools_and_none=True  # файл настроек
    )
    # theme = settings['GUI']['theme']
    # font_size = settings['FONT']['font_size']
    # # font_color = settings['FONT']['font_color']
    # font_family = settings['FONT']['font_family']
    # font_style = settings['FONT']['font_style']
    # path_font_size = settings['FONT']['path_font_size']
    # # path_font_color = settings['FONT']['path_font_color']
    # path_font_family = settings['FONT']['path_font_family']
    # path_font_style = settings['FONT']['path_font_style']
    # tab_font_size = settings['FONT']['tab_font_size']
    # # tab_font_color = settings['FONT']['tab_font_color']
    # tab_font_family = settings['FONT']['tab_font_family']
    # tab_font_style = settings['FONT']['tab_font_style']
    # sg.set_options(font=(font_family, font_size, font_style),
    #                tooltip_time=100, tooltip_font=(font_family, int(int(font_size)/1.2)))
    # print(sg.theme_text_color().upper())
    themes = (
        'без излишеств', 'раннее утро', 'ясный день', 'песчаный пляж', 'загорелый синий', 'морское дно', 'мурена',
        'искажённый лес', 'хвойный лес', 'зелёный чай', 'мятный макарон', 'яркие цвета', 'хот-дог', 'чирок-свистунок',
        'космос', 'лазурит', 'дождливый лес', 'дикий синий', 'васильки', 'лавандовый раф', 'песчаная буря', 'камин',
        'экокожа', 'кофе с молоком', 'графит с кремом', 'серая ночь', 'северное сияние', 'лунное затмение'
    )
    themes_dict = {
        'дикий синий': 'BlueMono',
        'васильки': 'BluePurple',
        'песчаная буря': 'DarkAmber',
        'лунное затмение': 'DarkBlack',
        'дождливый лес': 'DarkBlue11',
        'космос': 'DarkBlue17',
        'зелёный чай': 'DarkGreen',
        'графит с кремом': 'DarkGrey5',
        'серая ночь': 'DarkGrey11',
        'чирок-свистунок': 'DarkTeal',
        'искажённый лес': 'DarkTeal1',
        'морское дно': 'DarkTeal7',
        'мурена': 'DarkTeal8',
        'мятный макарон': 'GreenMono',
        'без излишеств': 'SystemDefault',
        'яркие цвета': 'LightGreen4',
        'ясный день': 'Material1',
        'лавандовый раф': 'Purple',
        'раннее утро': 'Reddit',
        'песчаный пляж': 'SandyBeach',
        'загорелый синий': 'TanBlue',
        'хот-дог': 'HotDogStand',
        'камин': 'Reds',
        'кофе с молоком': 'LightBrown11',
        'северное сияние': 'DarkTeal2',
        'лазурит': 'DarkBlue4',
        'экокожа': 'DarkBrown7',
        'хвойный лес': 'DarkGreen5'
    }
    tab_locations = {
        'сверху слева': 'topleft',
        'сверху посередине': 'top',
        'сверху справа': 'topright',
        'снизу слева': 'bottomleft',
        'снизу посередине': 'bottom',
        'снизу справа': 'bottomright',
    }
    font_styles = {
        'обычный': 'normal',
        'курсив': 'italic',
        'жирный': 'bold',
        'подчёркнутый': 'underline'
    }
    font_types = {
        'Основной текст': 'font',
        'Элементы меню': 'tab_font',
        'Текст путей': 'path_font'
    }

    # subsamples_dict = {
    #     '14': 39,
    #     '15': 38,
    #     '16': 37,
    #     '17': 36,
    #     '18': 35,
    #     '19': 34,
    #     '20': 30
    # }
    main_window()
